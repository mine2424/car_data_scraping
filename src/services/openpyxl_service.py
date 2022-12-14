from openpyxl import Workbook
from openpyxl import load_workbook


from constants.openpyxl_constants import OpenPyxlConstants


class OpenpyxlService:
    def __init__(self) -> None:
        self.sheet = None
        self.book = None
        self.excelFile = ''
        self.max_col = 1
        self.max_row = 2
        self.rows = None

    def add_max_col(self):
        self.max_col += 1

    def add_max_row(self):
        self.max_row += 1

    def clear_max_col_row(self):
        self.max_col = 0
        self.max_row = 0

    def init_openpyxl(self, fileName: str):
        initializedSheet = Workbook()
        self.excelFile = f'../data/{fileName}.xlsx'
        initializedSheet.save(self.excelFile)
        self.book = load_workbook(self.excelFile)
        self.sheet = self.book.worksheets[0]
        return self.sheet

    def openpyxl(self, fileName: str, sheet_index: int):
        if sheet_index is None:
            print('sheet_indexを入力してください')
            return
        self.excelFile = f'../data/{fileName}.xlsx'
        self.book = load_workbook(self.excelFile)
        self.sheet = self.book.worksheets[sheet_index]
        return self.sheet

    def save_sheet(self, fileName: str = ''):
        if fileName == '':
            self.book.save(self.excelFile)
        else:
            self.book.save(filename=fileName)

    def create_title(self):
        for i, indexTitle in enumerate(OpenPyxlConstants.titles):
            self.sheet.cell(row=1, column=i+1).value = indexTitle
        self.save_sheet()

    def create_used_car_title(self):
        for i, indexTitle in enumerate(OpenPyxlConstants.used_car_titles):
            self.sheet.cell(row=1, column=i+1).value = indexTitle
        self.save_sheet()

    def add_data_in_sheet(self, light_car_detail_dict: dict):
        for i, light_car_detail in enumerate(light_car_detail_dict.values()):
            self.sheet.cell(
                row=self.max_row, column=i + 1
            ).value = light_car_detail

        self.add_max_row()
        self.save_sheet()

    def add_used_car_data_in_sheet(self, val_list: list):
        for i, light_car_detail in enumerate(val_list):
            self.sheet.cell(
                row=self.max_row, column=i + 1
            ).value = light_car_detail

        self.add_max_row()
        self.save_sheet()

    def add_data_in_exit_file(self, col_index: int, row_index: int, val):
        self.sheet.cell(
            row=row_index, column=col_index
        ).value = val
        self.save_sheet()

    def remove_space_col(self):
        self.sheet.delete_cols(5)
        self.save_sheet()

    def get_colored_cell_index_list(self):
        index_list = []
        self.rows = self.sheet.rows
        for row in self.rows:
            is_colored = False
            for val in row:
                if 'FF92D050' == val.fill.fgColor.rgb:
                    is_colored = True
            if is_colored:
                # print(row[0].row)
                index_list.append(row[0].row)

        return index_list

    def get_max_row(self):
        self.max_row = self.sheet.max_row
